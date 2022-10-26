from discord import Webhook
import pandas as pd #pandas untuk membuat dataframe
from openpyxl import load_workbook #openpyxl untuk interaksi antara python dan file excel
from openpyxl.styles import *
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.label import DataLabelList
import string

input_file = 'Project/supermarket_sales.xlsx'
output_file = 'Output_penjualan_2019.xlsx'
# output_file = 'data_set/report_2019.csv'
sheetname = 'Report'
webhook_url = 'https://discord.com/api/webhooks/1026362803642187826/sOdbaZ2hm9zG7kR-HTSx6tEQxE8KKK-KovpabobLZWhVM21uX55VQmOGnws6RObCfazg'


df = pd.read_excel(input_file)

df_pivot = df.pivot_table(index='Gender', columns='Product line', values='Total', aggfunc='sum').round(decimals=2)


df_pivot.to_excel(excel_writer=output_file, sheet_name=sheetname)


#grafik
wb = load_workbook(output_file)
wb.active = wb[sheetname]

min_col = wb.active.min_column
max_col = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

# print(min_column, max_column, min_row, max_row)

barchart = BarChart()

data = Reference(wb.active, 
                    min_col=min_col+1, 
                    max_col=max_col, 
                    min_row=min_row, 
                    max_row=max_row
                )

categories = Reference(wb.active, 
                        min_col=min_col, 
                        max_col=max_col, 
                        min_row=min_row+1, 
                        max_row=max_row
                        )

barchart.add_data(data, titles_from_data=True)
barchart.set_categories(categories)

wb.active.add_chart(barchart, 'B6')
barchart.title = 'Sales berdasarkan produk'
barchart.style = 2

#Sum Penjualan
alphabet = list(string.ascii_uppercase)
alphabet_excel = alphabet[:max_col]

for i in alphabet_excel:
    if i != 'A':
        wb.active[f'{i}{max_row+1}'] = f'=SUM({i}{min_row+1}:{i}{max_row})'
        wb.active[f'{i}{max_row+1}'].style = 'Currency'

wb.active[f'{alphabet_excel[0]}{max_row+1}'] = 'Total'

wb.active['A1'] = 'Sales Report'
wb.active['A2'] = '2019'
wb.active['A1'].font = Font('Arial',bold=True,size=20)
wb.active['A1'].font = Font('Arial',bold=True,size=16)

wb.save(output_file)



def send_report_to_discord(webhookUrl, output):
    import discord
    from discord import SyncWebhook

    webhook = SyncWebhook.from_url(webhookUrl)

    with open(file=output, mode='rb') as file:
        excel_file = discord.File(file)
    
    webhook.send('This is an automated report', username='Data.IO', file=excel_file)


send_report_to_discord(webhook_url, output_file)